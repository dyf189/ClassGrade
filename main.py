import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import glob
import warnings
import os

# 忽略openpyxl图表相关的警告
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# 额外添加针对ExternalData.id的特定警告过滤
warnings.filterwarnings('ignore', message='.*ExternalData.*id should be.*', category=UserWarning)

class GradeSummaryGenerator:
    def __init__(self):
        # 科目定义
        self.grade7_subjects = ['语文', '数学', '英语', '地理', '生物', '历史', '政治']
        self.grade8_subjects = ['语文', '数学', '英语', '地理', '生物', '历史', '政治', '物理', '化学']
        
        # 样式定义
        self.header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        self.header_font = Font(bold=True, color="000000")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        
    def load_all_grades(self, data_dir):
        """加载所有成绩文件（支持.xls和.xlsx格式）"""
        all_files = []
        # 获取所有.xlsx文件
        all_files.extend(glob.glob(os.path.join(data_dir, "*.xlsx")))
        # 获取所有.xls文件
        all_files.extend(glob.glob(os.path.join(data_dir, "*.xls")))
        
        all_data = {}
        
        # 定义8个学期的完整列表，确保能识别所有学期
        semester_keys = ['初一上期中', '初一上期末', '初一下期中', '初一下期末', 
                         '初二上期中', '初二上期末', '初二下期中', '初二下期末', '初三上期中']
        
        for file in all_files:
            try:
                # 从文件名中提取学期和考试类型信息
                filename = os.path.basename(file)
                grade = None
                semester = None
                exam_type = None
                
                # 首先尝试精确匹配完整的学期键
                matched = False
                for key in semester_keys:
                    if key in filename:
                        grade = key[:2]  # 提取年级
                        semester = key[2:3]  # 提取学期(上/下)
                        exam_type = key[3:]  # 提取考试类型(期中/期末)
                        matched = True
                        break
                
                if not matched:
                    # 尝试传统的分步提取方式
                    # 提取年级信息
                    if "初一" in filename:
                        grade = "初一"
                    elif "初二" in filename:
                        grade = "初二"
                    elif "初三" in filename:
                        grade = "初三"
                    
                    # 提取学期信息
                    if "上" in filename:
                        semester = "上"
                    elif "下" in filename:
                        semester = "下"
                    
                    # 提取考试类型信息
                    if "期中" in filename:
                        exam_type = "期中"
                    elif "期末" in filename:
                        exam_type = "期末"
                
                if not grade or not semester or not exam_type:
                    print(f"无法从文件名 {filename} 中提取完整信息，跳过此文件")
                    continue
                
                key = f"{grade}{semester}{exam_type}"
                
                # 尝试读取Excel文件
                try:
                    df = pd.read_excel(file)
                    all_data[key] = df
                    print(f"成功加载: {key} - {filename}")
                except ImportError as e:
                    if 'xlrd' in str(e):
                        print(f"加载文件 {file} 时出错: 需要安装xlrd库来读取.xls文件。请运行: pip install xlrd")
                    else:
                        print(f"加载文件 {file} 时出错: {str(e)}")
                except Exception as e:
                    print(f"加载文件 {file} 时出错: {str(e)}")
            except Exception as e:
                print(f"处理文件 {file} 时出错: {str(e)}")
        
        # 打印加载的所有学期信息，方便调试
        print(f"成功加载的学期数据: {list(all_data.keys())}")
        return all_data
    
    def get_student_names(self, all_data):
        """从所有数据中提取所有学生姓名"""
        all_students = set()
        
        for df_key, df in all_data.items():
            try:
                # 尝试多种可能的姓名列名
                name_column_found = False
                
                # 尝试直接匹配'姓名'列
                if '姓名' in df.columns:
                    student_names = df['姓名'].dropna().astype(str).tolist()
                    all_students.update([name.strip() for name in student_names if name.strip()])
                    name_column_found = True
                else:
                    # 尝试包含'姓名'的列
                    name_columns = [col for col in df.columns if '姓名' in str(col) or '名字' in str(col)]
                    if name_columns:
                        for col in name_columns:
                            try:
                                student_names = df[col].dropna().astype(str).tolist()
                                all_students.update([name.strip() for name in student_names if name.strip()])
                                name_column_found = True
                                break  # 找到一个合适的列就够了
                            except:
                                continue
                
                if not name_column_found:
                    print(f"在文件 {df_key} 中未找到包含学生姓名的列")
            except Exception as e:
                print(f"提取文件 {df_key} 中的学生姓名时出错: {str(e)}")
        
        return all_students
    
    def is_merged_cell(self, ws, cell):
        """检查单元格是否是合并单元格的一部分"""
        for merged_cell in ws.merged_cells.ranges:
            if cell.coordinate in merged_cell:
                # 如果是合并单元格，返回左上角的单元格坐标
                return merged_cell.start_cell.coordinate
        return False
    
    def safe_write_cell(self, ws, row, col, value):
        """安全地写入单元格，处理合并单元格的情况"""
        try:
            # 首先获取单元格对象
            target_cell = ws.cell(row=row, column=col)
            # 检查是否是合并单元格
            merged_start_cell = self.is_merged_cell(ws, target_cell)
            if merged_start_cell:
                # 如果是合并单元格，只写入左上角单元格
                ws[merged_start_cell] = value
            else:
                # 不是合并单元格，直接写入
                target_cell.value = value
            return True
        except Exception as e:
            # 捕获所有可能的错误，如只读单元格等
            print(f"写入单元格({row}, {col})时出错: {str(e)}")
            return False
    
    def create_student_report(self, student_name, all_data, template_path, output_dir):
        """为单个学生创建报告"""
        try:
            # 加载模板
            wb = load_workbook(template_path)
            ws = wb.active
            
            # 设置学生姓名 - 使用安全写入方法
            self.safe_write_cell(ws, 2, 2, student_name)  # B2单元格
            
            # 定义数据位置 - 从第3行开始填充数据（学期从B3开始向下分布）
            data_start_row = 3
            
            # 包含期中期末的所有学期
            semesters = ['初一上期中', '初一上期末', '初一下期中', '初一下期末', 
                         '初二上期中', '初二上期末', '初二下期中', '初二下期末', '初三上期中']
            
            # 核心改进：使用模板中的科目名称作为基准，建立精确的映射关系
            template_subjects = self._get_template_subjects(ws)
            print(f"从模板中识别的科目映射: {template_subjects}")
            
            for i, semester in enumerate(semesters):
                # 计算当前学期应该填充的行号
                current_row = data_start_row + i
                print(f"\n正在处理 {student_name} 的 {semester} 数据，填充到行 {current_row}")
                
                # 检查数据是否存在
                if semester not in all_data:
                    print(f"警告: 未找到 {semester} 的数据，将填充空行")
                    # 即使没有数据，也填写学期名称，确保位置正确
                    self.safe_write_cell(ws, current_row, 2, semester)  # B列
                    continue
                
                df = all_data[semester]
                student_data = None
                
                # 查找学生数据 - 增强匹配逻辑
                name_columns_to_try = []
                # 先检查是否有'姓名'列
                if '姓名' in df.columns:
                    name_columns_to_try.append('姓名')
                # 再检查包含'姓名'或'名字'的列
                name_columns_to_try.extend([col for col in df.columns if '姓名' in str(col) or '名字' in str(col)])
                
                # 如果没有找到明显的姓名列，尝试所有可能的字符串列
                if not name_columns_to_try:
                    for col in df.columns:
                        try:
                            # 检查该列是否包含字符串类型的数据
                            if df[col].dtype == 'object' or isinstance(df[col].iloc[0], str):
                                name_columns_to_try.append(col)
                        except:
                            continue
                
                print(f"尝试的姓名列: {name_columns_to_try}")
                
                for col in name_columns_to_try:
                    try:
                        # 增强匹配逻辑，使用更灵活的字符串比较
                        student_data = df[df[col].astype(str).str.strip().str.contains(student_name, case=False, na=False)]
                        if not student_data.empty:
                            print(f"在列 '{col}' 中找到学生 '{student_name}'")
                            break
                    except Exception as e:
                        print(f"在列 '{col}' 中查找时出错: {str(e)}")
                        continue
                
                if student_data is None or student_data.empty:
                    # 没有找到该学生的数据，也要填写学期名称
                    print(f"警告: 未找到 {student_name} 在 {semester} 中的数据")
                    self.safe_write_cell(ws, current_row, 2, semester)  # B列
                    continue
                
                row = student_data.iloc[0]
                
                # 填写学期 - 使用安全写入方法
                self.safe_write_cell(ws, current_row, 2, semester)  # B列
                
                # 核心改进：根据模板科目名称建立数据列映射
                data_columns_mapping = self._map_data_columns(df.columns, template_subjects.keys())
                print(f"数据列映射关系: {data_columns_mapping}")
                
                # 改进的科目填充逻辑 - 精确填充
                filled_subjects = 0
                for template_subject, col_idx in template_subjects.items():
                    # 检查该科目是否在数据列映射中
                    if template_subject in data_columns_mapping:
                        data_col = data_columns_mapping[template_subject]
                        try:
                            if pd.notna(row[data_col]):
                                # 确保写入的是数值类型
                                value = row[data_col]
                                try:
                                    value = float(value)
                                except:
                                    print(f"警告: {data_col} 的值 '{value}' 无法转换为数值")
                                # 使用安全写入方法
                                self.safe_write_cell(ws, current_row, col_idx, value)
                                filled_subjects += 1
                                print(f"成功填充: {template_subject} -> 列 '{data_col}' -> 单元格({current_row}, {col_idx})")
                        except Exception as e:
                            print(f"填充 {template_subject} 时出错: {str(e)}")
                    else:
                        print(f"警告: 数据中没有找到与模板科目 '{template_subject}' 对应的列")
                
                print(f"成功填充了 {filled_subjects} 个科目数据")
                
                # 填写校排名和班排名
                self._fill_rank_data(ws, current_row, row)
            
            # 应用样式
            self.apply_styles(ws, data_start_row, len(semesters))
            
            # 保存文件
            output_path = os.path.join(output_dir, f"{student_name}_成绩总结.xlsx")
            wb.save(output_path)
            return True
            
        except Exception as e:
            print(f"生成 {student_name} 的报告时出错: {str(e)}")
            return False
            
    def _get_template_subjects(self, ws):
        """从模板中提取科目名称及其对应的列索引，根据用户提供的模板结构：科目在C列2行横向排列"""
        template_subjects = {}
        
        # 关键修改：根据用户提供的信息，科目在第2行（row=2）从C列（column=3）开始横向排列
        header_row = 2  # 科目标题行固定为第2行
        start_col = 3   # 科目从第3列（C列）开始
        
        # 定义标准科目名称列表，用于验证
        standard_subjects = {'语文', '数学', '英语', '地理', '生物', '历史', '政治', '物理', '化学', '总分'}
        
        # 从第3列（C列）开始向右查找科目名称，最多检查到第15列
        for col_idx in range(start_col, 16):
            cell_value = ws.cell(row=header_row, column=col_idx).value
            
            # 检查单元格值是否有效
            if cell_value and isinstance(cell_value, str):
                # 清理科目名称
                clean_subject = self._clean_subject_name(cell_value)
                
                # 添加到模板科目映射中
                if clean_subject in standard_subjects:
                    template_subjects[clean_subject] = col_idx
                    print(f"识别科目: '{clean_subject}' 在单元格({header_row}, {col_idx})")
            # 如果单元格为空，且已经找到了至少一个科目，可能是科目列表结束
            elif template_subjects and col_idx > start_col:
                print(f"遇到空单元格，停止查找科目")
                break
        
        # 如果没有找到任何科目，使用默认映射
        if not template_subjects:
            print("未从模板中识别到科目，使用默认映射")
            # 默认映射：从第3列（C列）开始依次排列标准科目
            default_subjects = ['语文', '数学', '英语', '物理', '政治', '历史', '生物', '地理', '化学', '总分']
            template_subjects = {subject: start_col + i for i, subject in enumerate(default_subjects)}
        
        return template_subjects
        
    def _clean_subject_name(self, subject_name):
        """清理科目名称，移除前缀、后缀和特殊字符"""
        # 去除空格
        clean_name = subject_name.strip().replace(' ', '')
# 定义要移除的常见后缀
        suffixes = ['成绩', '分数', '得分', '分', '绩', '考试', '科目', '期中', '期末']
        for suffix in suffixes:
            if clean_name.endswith(suffix):
                clean_name = clean_name[:-len(suffix)]
                break
        
        # 定义要移除的常见前缀
        prefixes = ['期中', '期末', '上学期', '下学期', '初一', '初二', '初三']
        for prefix in prefixes:
            if clean_name.startswith(prefix):
                clean_name = clean_name[len(prefix):]
                break
        
        return clean_name

    def _map_data_columns(self, data_columns, template_subjects):
        """将数据列与模板科目进行精确映射"""
        mapping = {}
        
        # 为每个模板科目查找匹配的数据列
        for subject in template_subjects:
            # 创建匹配模式
            patterns = [
                subject,
                f"{subject}成绩", f"{subject}分数", f"{subject}得分", f"{subject}分",
                f"期中{subject}", f"期末{subject}", f"{subject}期中", f"{subject}期末"
            ]

            if subject == "总分":
                patterns.extend(["总分", "总分成绩", "总分分数", "总分得分", "总分分", "总分数", "总成绩"])
            
            # 查找匹配的列
            matched_column = self._find_matching_column(data_columns, patterns)
            if matched_column:
                mapping[subject] = matched_column
        
        return mapping
        
    def _find_matching_column(self, columns, patterns):
        """在列列表中查找与任何模式匹配的列"""
        # 首先尝试精确匹配
        for col in columns:
            col_str = str(col).strip()
            for pattern in patterns:
                if col_str == pattern or col_str.lower() == pattern.lower():
                    return col
        
        # 然后尝试包含匹配
        for col in columns:
            col_str = str(col).strip().lower()
            for pattern in patterns:
                if pattern.lower() in col_str:
                    return col
        
        return None
        
    def _fill_rank_data(self, ws, current_row, row_data):
        """填充排名数据"""
        # 定义排名类型和对应的列索引
        rank_types = {
            '校排名': 14,  # 校排名在第11列
            '班排名': 13   # 班排名在第12列
        }
        
        # 填充校排名
        school_rank_found = False
        school_rank_variants = ['校排名', '校名次', '序号', '校次']
        for variant in school_rank_variants:
            if variant in row_data.index:
                try:
                    if pd.notna(row_data[variant]):
                        value = row_data[variant]
                        try:
                            value = int(value)
                        except:
                            pass
                        self.safe_write_cell(ws, current_row, rank_types['校排名'], value)
                        print(f"填充校排名: 从列 '{variant}' 获取值 '{value}'")
                        school_rank_found = True
                        break
                except Exception as e:
                    print(f"填充校排名时出错: {str(e)}")
        
        if not school_rank_found:
            print("未找到校排名数据")
        
        # 填充班排名
        class_rank_found = False
        class_rank_variants = ['班排名', '班名次', '班级排名', '班级名次']
        for variant in class_rank_variants:
            if variant in row_data.index:
                try:
                    if pd.notna(row_data[variant]):
                        value = row_data[variant]
                        try:
                            value = int(value)
                        except:
                            pass
                        self.safe_write_cell(ws, current_row, rank_types['班排名'], value)
                        print(f"填充班排名: 从列 '{variant}' 获取值 '{value}'")
                        class_rank_found = True
                        break
                except Exception as e:
                    print(f"填充班排名时出错: {str(e)}")
        
        if not class_rank_found:
            print("未找到班排名数据")
            
            # 新增：填充总分数据（固定在第12列，L列）- 移到正确位置
        total_score_found = False
        total_score_variants = ['总分', '总分成绩', '总分分数', '总分得分', '总分分', '总分数', '总成绩']
            
        for variant in total_score_variants:
            if variant in row_data.index:
                try:
                    if pd.notna(row_data[variant]):
                        value = row_data[variant]
                        try:
                            value = float(value)
                        except:
                            pass
                            # 总分固定在第12列（L列）
                        self.safe_write_cell(ws, current_row, 12, value)
                        print(f"填充总分: 从列 '{variant}' 获取值 '{value}' -> 单元格({current_row}, 12)")
                        total_score_found = True
                        break
                except Exception as e:
                    print(f"填充总分时出错: {str(e)}")
            
        if not total_score_found:
            print("未找到总分数据")

    def apply_styles(self, ws, start_row, row_count):
        """应用样式到工作表，避免处理合并单元格"""
        # 创建一个合并单元格坐标的集合，用于快速检查
        merged_cells_coords = set()
        for merged_range in ws.merged_cells.ranges:
            # 获取合并区域内的所有单元格坐标
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    merged_cells_coords.add(cell.coordinate)
        
        # 关键修改：设置标题行样式 - 科目在第2行（row=2）
        for col in range(1, 16):  # 增加列范围以覆盖所有科目列
            cell = ws.cell(row=2, column=col)  # 标题行改为第2行
            # 跳过合并单元格
            if cell.coordinate not in merged_cells_coords:
                # 移除cell.fill = self.header_fill 这一行，取消黄色填充
                cell.font = self.header_font
                cell.border = self.border
                cell.alignment = self.center_alignment
        
        # 设置数据区域样式
        for row in range(start_row, start_row + row_count):
            for col in range(1, 16):  # 增加列范围以覆盖所有科目列
                cell = ws.cell(row=row, column=col)
                # 跳过合并单元格
                if cell.coordinate not in merged_cells_coords:
                    try:
                        cell.border = self.border
                        cell.alignment = self.center_alignment
                    except:
                        # 忽略样式应用错误
                        pass
    
    def generate_all_reports(self, data_dir, template_path, output_dir):
        """生成所有学生的报告"""
        # 创建输出目录
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # 加载所有成绩数据
        all_data = self.load_all_grades(data_dir)
        
        if not all_data:
            print("未找到任何成绩文件!")
            return
        
        # 获取所有学生姓名
        all_students = self.get_student_names(all_data)
        
        if not all_students:
            print("未找到任何学生姓名，请检查成绩文件格式")
            return
        
        print(f"找到 {len(all_students)} 名学生")
        
        # 为每名学生生成报告
        success_count = 0
        for student in all_students:
            if self.create_student_report(student, all_data, template_path, output_dir):
                success_count += 1
                if success_count % 10 == 0:
                    print(f"已生成 {success_count} 份报告...")
        
        print(f"成功生成 {success_count} 份学生成绩报告，保存在 {output_dir} 目录中")

class GradeBefore200:
    def __init__(self):
        # 样式定义
        self.header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        self.header_font = Font(bold=True, color="000000")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        
        # 定义9个学期
        self.semesters = ['初一上期中', '初一上期末', '初一下期中', '初一下期末', 
                         '初二上期中', '初二上期末', '初二下期中', '初二下期末', '初三上期中']
    
    def create_template_file(self, template_path):
        """创建前200名统计模板文件"""
        try:
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            
            # 设置表头
            ws['A1'] = "学期统计"
            ws['A2'] = "学期"
            ws['A3'] = "人数"
            ws['A4'] = "学生姓名"
            
            # 设置样式
            ws['A1'].font = self.header_font
            ws['A2'].font = self.header_font
            ws['A3'].font = self.header_font
            ws['A4'].font = self.header_font
            
            # 保存模板文件
            wb.save(template_path)
            print(f"已创建新的模板文件: {template_path}")
            return True
        except Exception as e:
            print(f"创建模板文件时出错: {str(e)}")
            return False
    
    def load_all_grades(self, data_dir):
        """加载所有成绩文件（复用GradeSummaryGenerator的加载逻辑）"""
        all_files = []
        all_files.extend(glob.glob(os.path.join(data_dir, "*.xlsx")))
        all_files.extend(glob.glob(os.path.join(data_dir, "*.xls")))
        
        all_data = {}
        
        for file in all_files:
            try:
                filename = os.path.basename(file)
                grade = None
                semester = None
                exam_type = None
                
                # 匹配学期
                matched = False
                for key in self.semesters:
                    if key in filename:
                        grade = key[:2]
                        semester = key[2:3]
                        exam_type = key[3:]
                        matched = True
                        break
                
                if not matched:
                    if "初一" in filename:
                        grade = "初一"
                    elif "初二" in filename:
                        grade = "初二"
                    elif "初三" in filename:
                        grade = "初三"
                    
                    if "上" in filename:
                        semester = "上"
                    elif "下" in filename:
                        semester = "下"
                    
                    if "期中" in filename:
                        exam_type = "期中"
                    elif "期末" in filename:
                        exam_type = "期末"
                
                if not grade or not semester or not exam_type:
                    print(f"无法从文件名 {filename} 中提取完整信息，跳过此文件")
                    continue
                
                key = f"{grade}{semester}{exam_type}"
                
                try:
                    df = pd.read_excel(file)
                    all_data[key] = df
                    print(f"成功加载: {key} - {filename}")
                except Exception as e:
                    print(f"加载文件 {file} 时出错: {str(e)}")
            except Exception as e:
                print(f"处理文件 {file} 时出错: {str(e)}")
        
        return all_data
    
    def find_school_rank_column(self, df):
        """查找校排名列"""
        rank_columns = ['校排名', '校名次', '序号', '校次', '名次']
        for col in df.columns:
            col_str = str(col).strip()
            for rank_col in rank_columns:
                if rank_col in col_str:
                    return col
        return None
    
    def find_student_name_column(self, df):
        """查找学生姓名列"""
        name_columns = ['姓名', '名字', '学生姓名', '学生名字']
        for col in df.columns:
            col_str = str(col).strip()
            for name_col in name_columns:
                if name_col in col_str:
                    return col
        return None
    
    def get_top_200_students(self, df):
        """获取前200名学生（校排名<=200）"""
        rank_col = self.find_school_rank_column(df)
        name_col = self.find_student_name_column(df)
        
        if not rank_col or not name_col:
            print(f"未找到排名列或姓名列，跳过此文件")
            return []
        
        try:
            # 过滤出校排名<=200的学生
            df_filtered = df[pd.to_numeric(df[rank_col], errors='coerce') <= 200]
            
            # 按排名排序
            df_sorted = df_filtered.sort_values(by=rank_col, ascending=True)
            
            # 提取学生姓名
            top_students = df_sorted[name_col].dropna().astype(str).tolist()
            
            return top_students
        except Exception as e:
            print(f"获取前200名学生时出错: {str(e)}")
            return []
    
    def safe_write_cell(self, ws, row, col, value):
        """安全地写入单元格，处理合并单元格的情况"""
        try:
            target_cell = ws.cell(row=row, column=col)
            # 检查是否是合并单元格
            for merged_cell in ws.merged_cells.ranges:
                if target_cell.coordinate in merged_cell:
                    # 如果是合并单元格，写入左上角单元格
                    ws[merged_cell.start_cell.coordinate] = value
                    return True
            # 不是合并单元格，直接写入
            target_cell.value = value
            return True
        except Exception as e:
            print(f"写入单元格({row}, {col})时出错: {str(e)}")
            return False
    
    def apply_styles(self, ws):
        """应用样式到工作表"""
        # 设置标题行样式（第2行和第3行）
        for row in [2, 3]:
            for col in range(2, 12):  # B列到K列
                cell = ws.cell(row=row, column=col)
                cell.font = self.header_font
                cell.border = self.border
                cell.alignment = self.center_alignment
                if row == 2:  # 第2行添加黄色填充
                    cell.fill = self.header_fill
        
        # 设置学生姓名区域样式（从第4行开始）
        max_row = ws.max_row
        for row in range(4, max_row + 1):
            for col in range(2, 12):  # B列到K列
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
                cell.alignment = self.center_alignment
    
    def generate_top200_report(self, data_dir, template_path, output_path):
        """生成前200名统计报告"""
        try:
            # 检查模板文件是否存在且有效
            if not os.path.exists(template_path):
                print(f"模板文件 '{template_path}' 不存在，尝试创建新模板...")
                if not self.create_template_file(template_path):
                    print("创建模板文件失败，无法生成报告")
                    return False
            
            # 尝试加载模板
            try:
                wb = load_workbook(template_path)
                ws = wb.active
                print(f"成功加载模板文件: {template_path}")
            except Exception as e:
                print(f"加载模板文件 '{template_path}' 时出错: {str(e)}")
                print("尝试创建新的模板文件...")
                if not self.create_template_file(template_path):
                    print("创建模板文件失败，无法生成报告")
                    return False
                # 重新加载新创建的模板
                wb = load_workbook(template_path)
                ws = wb.active
            
            # 加载成绩数据
            all_data = self.load_all_grades(data_dir)
            
            if not all_data:
                print("未找到任何成绩文件!")
                return False
            
            print(f"成功加载 {len(all_data)} 个学期的数据")
            
            # 填充学期名称（B2往右）
            for i, semester in enumerate(self.semesters):
                col = 2 + i  # B列开始，依次向右
                self.safe_write_cell(ws, 2, col, semester)
                print(f"填充学期名称: {semester} -> 单元格(2, {col})")
            
            # 统计每个学期的前200名学生并填充
            for i, semester in enumerate(self.semesters):
                col = 2 + i  # B列开始，依次向右
                
                if semester not in all_data:
                    print(f"警告: 未找到 {semester} 的数据")
                    # 填充人数为0
                    self.safe_write_cell(ws, 3, col, 0)
                    continue
                
                df = all_data[semester]
                top_students = self.get_top_200_students(df)
                
                # 填充人数（B3往右）
                self.safe_write_cell(ws, 3, col, len(top_students))
                print(f"填充人数: {semester} -> {len(top_students)}人 -> 单元格(3, {col})")
                
                # 填充学生姓名（从B4开始往下）
                for j, student_name in enumerate(top_students):
                    row = 4 + j  # 从第4行开始
                    self.safe_write_cell(ws, row, col, student_name)
                    if j < 5:  # 只打印前5个学生的填充信息，避免输出过多
                        print(f"填充学生姓名: {student_name} -> 单元格({row}, {col})")
                
                if len(top_students) > 5:
                    print(f"... 共填充 {len(top_students)} 名学生")
            
            # 应用样式
            self.apply_styles(ws)
            
            # 保存文件
            wb.save(output_path)
            print(f"前200名统计报告已生成: {output_path}")
            return True
            
        except Exception as e:
            print(f"生成前200名统计报告时出错: {str(e)}")
            return False

# 使用示例
if __name__ == "__main__":
    # 提示安装xlrd库（如果需要）
    try:
        import xlrd
    except ImportError:
        print("提示: 检测到未安装xlrd库，将无法读取.xls格式的成绩文件。")
        print("请运行以下命令安装xlrd库：")
        print("pip install xlrd")
        print("\n如果您只需要处理.xlsx格式的文件，可以忽略此提示。")
        print("")
    
    generator = GradeSummaryGenerator()
    
    # 设置路径
    data_directory = "./Grade/"  # 存放8个成绩文件的目录
    template_file = "成绩模板.xlsx"  # Excel模板文件
    output_directory = "学生成绩报告"  # 输出目录
    
    # 生成所有报告
    generator.generate_all_reports(data_directory, template_file, output_directory)
    
    # 新增：生成前200名统计报告
    top200_generator = GradeBefore200()
    top200_template = "前200名.xlsx"  # 前200名模板文件
    top200_output = "全校前200名统计.xlsx"  # 前200名统计输出文件
    
    # 检查模板文件是否存在，如果不存在或损坏则创建新模板
    if not os.path.exists(top200_template):
        print(f"前200名模板文件 '{top200_template}' 不存在，将创建新模板")
        top200_generator.create_template_file(top200_template)
    
    top200_generator.generate_top200_report(data_directory, top200_template, top200_output)